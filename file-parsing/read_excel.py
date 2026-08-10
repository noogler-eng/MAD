import openpyxl

wb = openpyxl.load_workbook("students.xlsx", data_only=True)
ws = wb["Students"]

print(f"Reading data from Excel file 'students.xlsx': {ws.max_row} rows, {ws.max_column} columns")

# skipping 1st row as it contains headers
# values_only=True will give us the values of the cells instead of cell objects
for row in ws.iter_rows(min_row=2, values_only=True):
    name, age, grade = row
    print(f"Name: {name}, Age: {age}, Grade: {grade}")